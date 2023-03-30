import os

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

PATH = 'media/temp/'
RESPONSE_NAME = 'Отчет.xlsx'
FULL_PATH = PATH + RESPONSE_NAME
SAMPLE = 'media/Шапка отчета.xlsx'


def handle_file(file):
    """Организация обработки файла"""

    # создание таблицы
    df = create_df(file)
    # рассчет значений для столбца "Исчислено всего по формуле"
    df['Исчислено всего по формуле'] = df.apply(calculate_total, axis=1)
    # рассчет значений для столбца "Отклонения"
    df['Отклонения'] = df['Исчислено всего'] - df['Исчислено всего по формуле']
    df.dropna(subset=['Филиал'], inplace=True)
    # сортировка таблицы
    df.sort_values(by='Отклонения', ascending=False, inplace=True)
    # отправка таблицы
    return response_table(df)


def create_df(file) -> pd.DataFrame:
    """извлечение данных из исходного df"""
    source_df = pd.read_excel(file)
    df = pd.DataFrame()

    # поиск и заполнение данными из исходника
    header = ['Филиал', 'Сотрудник', 'Налоговая база', 'Исчислено всего']
    for h in header:
        new_name = None
        new_col = None
        for column in source_df.items():
            name, col = column
            if h == name:
                new_name = name
                new_col = col
                break
            elif col[0] and col[0] == h:
                new_name = col[0]
                new_col = col
                break
        if new_name:
            df[new_name] = new_col
            new_name = None

    return df


def calculate_total(row: pd.Series):
    """рассчет значений для столбца 'Исчислено всего по формуле'"""
    if row['Налоговая база'] < 5000000:
        return row['Налоговая база'] * 0.13
    else:
        return row['Налоговая база'] * 0.15


def response_table(df: pd.DataFrame):
    """возврат результата обработки файла"""
    # создание нового Excel файла
    wb = load_workbook(SAMPLE)
    ws = wb.active

    # добавление данных в Excel файл
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # форматирование ячеек столбца "Отклонения"
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            if cell.value == 0:
                cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # добавление директории
    if not os.path.isdir(PATH):
        os.makedirs(PATH)
    # очистка директории
    if os.path.exists(FULL_PATH):
        os.remove(FULL_PATH)

    # сохранение отчета в новом Excel файле
    wb.save(FULL_PATH)
    return FULL_PATH
